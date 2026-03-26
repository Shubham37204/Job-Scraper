import smtplib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import date
import os

SENDER_EMAIL = "your email here"
SENDER_PASSWORD = "gmail app generated password not your normal gmail account password"
RECEIVER_EMAIL = "your email here"

KEYWORD_MAP = {
    "python":      "Python, OOP, FastAPI, Django, REST API, Flask",
    "django":      "Django, REST Framework, ORM, PostgreSQL, Python",
    "flask":       "Flask, REST API, Python, Microservices",
    "full stack":  "React, Node.js, REST API, SQL, Git, HTML/CSS",
    "frontend":    "React, Next.js, HTML, CSS, JavaScript, TypeScript",
    "backend":     "Node.js, Express, REST API, SQL, Docker, Git",
    "next":        "Next.js, React, TypeScript, Vercel, SSR, SEO",
    "react":       "React, Redux, TypeScript, REST API, Webpack",
    "aws":         "AWS EC2, S3, Lambda, IAM, CloudFormation, Terraform",
    "cloud":       "AWS/GCP/Azure, Docker, Kubernetes, CI/CD, Terraform",
    "ai":          "Python, TensorFlow/PyTorch, LLMs, Prompt Engineering, ML",
    "ml":          "Python, Scikit-learn, TensorFlow, Pandas, NumPy, MLOps",
    "data":        "SQL, Python, Pandas, Power BI/Tableau, ETL, Statistics",
    "devops":      "Docker, Kubernetes, CI/CD, Terraform, Linux, Jenkins",
    "mobile":      "Flutter/React Native, iOS/Android, REST API, Firebase",
    "java":        "Java, Spring Boot, Microservices, Maven, REST API",
    "node":        "Node.js, Express, MongoDB, REST API, TypeScript",
}


def get_keywords(title):
    """Match job title against keyword map and return relevant resume keywords"""
    title_lower = title.lower()
    matched = []
    for trigger, keywords in KEYWORD_MAP.items():
        if trigger in title_lower:
            # Avoid duplicate keyword sets
            if keywords not in matched:
                matched.append(keywords)
    return " | ".join(matched) if matched else "Python, Git, REST API, SQL"


def is_recent(posted):
    """Returns True only if job was posted recently — filters out 2/3 weeks ago"""
    if not posted or posted == "N/A":
        return True   
    posted_lower = posted.lower()

    recent_signals = ["today", "hour", "just", "1 day", "2 day",
                      "3 day", "4 day", "5 day", "6 day", "7 day", "1 week"]
    for signal in recent_signals:
        if signal in posted_lower:
            return True

    old_signals = ["2 week", "3 week", "4 week", "month", "30", "ago"]
    for signal in old_signals:
        if signal in posted_lower:
            return False

    return True  

def style_sheet(ws, jobs, source_color):
    """Writes jobs to a worksheet with full formatting"""
    headers = ["#", "Title", "Company", "Posted", "Resume Keywords", "Link"]

    header_fill = PatternFill("solid", fgColor=source_color)
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for idx, job in enumerate(jobs, start=1):
        keywords = get_keywords(job.get("title", ""))
        ws.append([
            idx,
            job.get("title",   "N/A"),
            job.get("company", "N/A"),
            job.get("posted",  "N/A"),
            keywords,
            job.get("link",    "N/A"),
        ])

        link_cell = ws.cell(row=idx + 1, column=6)
        link_cell.hyperlink = job.get("link", "")
        link_cell.font = Font(color="0000FF", underline="single")

        if idx % 2 == 0:
            for col in range(1, 7):
                ws.cell(
                    row=idx + 1, column=col).fill = PatternFill("solid", fgColor="F2F2F2")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 55


def create_excel(jobs):
    """Creates Excel with one sheet per source + one summary sheet"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  

    source_colors = {
        "Internshala": "E87722",   
        "Indeed":      "2164F3",   
        "Wellfound":   "EE4444",  
        "All Jobs":    "4CAF50",   
    }

    sources = sorted(set(job["source"] for job in jobs))
    for source in sources:
        source_jobs = [j for j in jobs if j["source"] == source]
        ws = wb.create_sheet(title=source)
        color = source_colors.get(source, "888888")
        ws.sheet_properties.tabColor = color
        style_sheet(ws, source_jobs, color)


    ws_all = wb.create_sheet(title="All Jobs", index=0) 
    ws_all.sheet_properties.tabColor = source_colors["All Jobs"]
    style_sheet(ws_all, jobs, source_colors["All Jobs"])

    filename = f"jobs_{date.today()}.xlsx"
    wb.save(filename)
    return filename


def send_digest_email(jobs):
    if not jobs:
        print("📭 No new jobs to send.")
        return

    today = date.today().strftime("%d %B %Y")
    sources = {}
    for job in jobs:
        sources.setdefault(job["source"], 0)
        sources[job["source"]] += 1

    breakdown = " | ".join(f"{src}: {cnt}" for src, cnt in sources.items())

    th = "padding:10px; border:1px solid #ddd; text-align:left; background:#4CAF50; color:white; font-size:13px;"
    td = "padding:8px 10px; border:1px solid #ddd; font-size:12px;"

    rows_html = ""
    for idx, job in enumerate(jobs, start=1):
        bg = "#f9f9f9" if idx % 2 == 0 else "#ffffff"
        keywords = get_keywords(job.get("title", ""))
        rows_html += f"""
        <tr style="background:{bg};">
            <td style="{td}">{idx}</td>
            <td style="{td}">{job.get('title', 'N/A')}</td>
            <td style="{td}">{job.get('company', 'N/A')}</td>
            <td style="{td}">{job.get('source', 'N/A')}</td>
            <td style="{td}">{job.get('posted', 'N/A')}</td>
            <td style="{td}; font-size:11px; color:#555;">{keywords}</td>
            <td style="{td}">
                <a href="{job.get('link', '')}"
                   style="background:#4CAF50; color:white; padding:5px 12px;
                          border-radius:4px; text-decoration:none; font-size:12px;">
                    Apply
                </a>
            </td>
        </tr>"""

    html = f"""
    <div style="font-family:Arial,sans-serif; max-width:1000px; margin:auto;">
        <h2 style="color:#4CAF50;">🆕 Job Digest — {today}</h2>
        <p style="color:#555;">{len(jobs)} new job(s) | {breakdown}</p>
        <table style="width:100%; border-collapse:collapse;">
            <thead>
                <tr>
                    <th style="{th}">#</th>
                    <th style="{th}">Title</th>
                    <th style="{th}">Company</th>
                    <th style="{th}">Source</th>
                    <th style="{th}">Posted</th>
                    <th style="{th}">Resume Keywords</th>
                    <th style="{th}">Apply</th>
                </tr>
            </thead>
            <tbody>{rows_html}</tbody>
        </table>
        <p style="color:#aaa; font-size:11px; margin-top:20px;">
            Sent by your Job Scraper Bot 🤖 | Excel attached with separate tabs per source
        </p>
    </div>
    """

    msg = MIMEMultipart("mixed")
    msg["Subject"] = f"🆕 {len(jobs)} New Jobs — {today} | {breakdown}"
    msg["From"] = SENDER_EMAIL
    msg["To"] = RECEIVER_EMAIL
    msg.attach(MIMEText(html, "html"))

    excel_path = create_excel(jobs)
    with open(excel_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition",
                        f"attachment; filename={excel_path}")
        msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, RECEIVER_EMAIL, msg.as_string())

    os.remove(excel_path)
    print(f"📧 Digest sent — {len(jobs)} jobs | {breakdown}")
